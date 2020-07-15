from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt 
from sklearn import preprocessing
import numpy as np

wb = load_workbook(filename=r'2009_YueXiu.xlsx')

sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0])

data = []


for rx in range(1, ws.max_row + 1):
    temp_list = []
    pid = rx
    w1 = ws.cell(row=rx, column=1).value
    w2 = ws.cell(row=rx, column=2).value
    w3 = ws.cell(row=rx, column=3).value
    temp_list.append(w1)
    temp_list.append(w2)
    temp_list.append(w3)
    
    data.append(temp_list)

print(len(data))

pic = []
for i in range(len(data)):
    pic.append(data[i][2])

x = range(0, 53)
y = pic
plt.plot(x, y, color='r',marker='o', markerfacecolor='blue',markersize=5)
plt.xlabel('Week Number') 
plt.ylabel('Case Number') 
plt.title('Tendency') 
plt.legend() 
plt.show() 

pic_scaled = preprocessing.scale(pic)
temp = np.array(pic_scaled) 
x = range(0, 53)
y = pic_scaled
plt.plot(x, y, color='r',marker='o', markerfacecolor='blue',markersize=5)
plt.xlabel('Week Number') 
plt.ylabel('Case Number') 
plt.title('Tendency') 
plt.legend() 
plt.show() 