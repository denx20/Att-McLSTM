from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt

wb = load_workbook(filename=r'2017_raw.xlsx')
ws = wb.active 

#load data
temp1 = []
for row in ws.rows:
    temp2 = []
    for col in row:
        temp2.append(col.value)
    temp1.append(temp2)
data = temp1[1:]

#get col
def get_col(col_num):
    col = []
    for i in range(len(data)):
        col.append(data[i][col_num])
    return col

#calculate city total value
ss = []
for i in range(int(len(data)/11)):
    j = i * 11
    k = i * 11 + 10
    t = (j, k)
    ss.append(t)
temp = []
for i in ss:
    
    temp1 = data[i[0]:(i[1] + 1)]
    temp2 = [0] * 49
    temp2[0] = 0
    temp2[1] = temp1[0][1]
    temp2[2] = temp1[0][2]
    temp2[3] = temp1[0][3]
    temp2[4] = 'Guang Zhou'
    for i in range(5, 41):
        s = 0
        for j in range(len(temp1)):
            s = s + temp1[j][i]
        temp2[i] = s
        #temp2[5:40]
    temp2[41] = temp1[0][41]
    temp2[42] = temp1[0][42]
    temp2[43] = temp1[0][43]
    temp2[44] = temp1[0][44]
    temp2[45] = temp1[0][45]
    temp2[46] = temp1[0][46]
    temp2[47] = temp1[0][47]
    temp2[48] = temp1[0][48]
    temp.append(temp2)

#White new data
wb = Workbook()
ws = wb.active
for row1 in range(1, len(temp) + 1):
    for col1 in range(1, len(temp[0]) + 1):
        ws.cell(row = row1, column = col1, value = temp[row1 - 1][col1 - 1])
wb.save('2017_sum.xlsx')
