from keras.models import Sequential
from keras.layers import Dense, LSTM
from keras.optimizers import SGD
import numpy as np
import matplotlib.pyplot as plt
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from sklearn import preprocessing
import tensorflow as tf
from tensorflow.contrib import rnn
from tensorflow.contrib.learn.python.learn.estimators.estimator import SKCompat
from matplotlib import style
import pandas as pd

wb = load_workbook(filename=r'2009_YueXiu.xlsx')

sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0])

data = []


for rx in range(1, ws.max_row + 1):
    temp_list = []
    pid = rx
    w1 = ws.cell(row=rx, column=1).value
    w2 = ws.cell(row=rx, column=2).value
    w3 = ws.cell(row=rx, column=3).value
    temp_list.append(w1)
    temp_list.append(w2)
    temp_list.append(w3)
    
    data.append(temp_list)

print(len(data))

pic = []
for i in range(len(data)):
    pic.append(data[i][2])
pic_scaled = preprocessing.scale(pic)
temp = np.array(pic_scaled) 

def generate_data(seq):
    X = []
    Y = []

    for i in range(len(seq) - TIMESTEPS - 1):
        X.append([seq[i:i + TIMESTEPS]])
        Y.append([seq[i + TIMESTEPS]])
    return np.array(X, dtype=np.float32), np.array(Y, dtype=np.float32)

HIDDEN_SIZE = 40
NUM_LAYERS = 3
TIMESTEPS = 20
TRAINING_STEPS = 100
BATCH_SIZE = 20

def LstmCell():
    lstm_cell = rnn.BasicLSTMCell(HIDDEN_SIZE, state_is_tuple=True)#
    return lstm_cell

def lstm_model(X, y):
    cell = rnn.MultiRNNCell([LstmCell() for _ in range(NUM_LAYERS)])
    output, _ = tf.nn.dynamic_rnn(cell, X, dtype=tf.float32)
    output = tf.reshape(output, [-1, HIDDEN_SIZE])
    predictions = tf.contrib.layers.fully_connected(output, 1, None)
    labels = tf.reshape(y, [-1])
    predictions = tf.reshape(predictions, [-1])
    loss = tf.losses.mean_squared_error(predictions, labels)
    train_op = tf.contrib.layers.optimize_loss(loss,
                                               tf.contrib.framework.get_global_step(),
                                               optimizer='Adagrad',
                                               learning_rate=0.6)
    return predictions, loss, train_op

learn = tf.contrib.learn
regressor = SKCompat(learn.Estimator(model_fn=lstm_model, model_dir='Models/model_2'))
train_X, train_y = generate_data(temp)
test_X, test_y = generate_data(temp)
regressor.fit(train_X, train_y, batch_size=BATCH_SIZE, steps=TRAINING_STEPS)
predicted = np.array([pred for pred in regressor.predict(test_X)])

'''绘制反标准化之前的真实值与预测值对比图'''
plt.figure()
plt.plot(predicted, label='预测值')
plt.plot(test_y, label='真实值')
plt.title('反标准化之前')
plt.legend()
plt.show()

mape = np.mean(abs((predicted - test_y) / predicted)) * 100
print(mape)