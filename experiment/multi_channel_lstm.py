from math import *
import numpy as np
import os
from matplotlib import pyplot
import matplotlib.pyplot as plt
from pandas import read_csv
from pandas import DataFrame
from pandas import concat
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from sklearn.preprocessing import MinMaxScaler
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import mean_squared_error
from keras import regularizers
from keras.optimizers import Adam
from keras.utils import np_utils
from keras.models import *
from keras.models import Sequential, Model
from keras.layers import *
from keras.layers import Dense, Input, concatenate, LSTM, merge
from keras.layers.core import *
import random

#constant
TOTAL_WEEK = 468
TRAIN_WEEK = 370
WEEK_LENGTH = 6
TIME_STEPS = WEEK_LENGTH
AREA_NUM = 10

#########################################
#load content data
pre_XX = []
path = 'C:\\Users\\HP\\Desktop\\实验\\流感疫情分析\\工作进展\\2_LSTM预测\\4_lstm_for_structure2\\content'
files= os.listdir(path)
for file in files:
    filename = path + '\\' + file
    wb = load_workbook(filename = filename)
    ws = wb.active 
    temp1 = []
    for row in ws.rows:
        temp2 = []
        for col in row:
            temp2.append(col.value)
        temp1.append(temp2)
    temp2 = temp1[1:]
    pre_XX += temp2

#load weather data
pre_XX_wea = []
filename = 'C:\\Users\\HP\\Desktop\\实验\\流感疫情分析\\工作进展\\2_LSTM预测\\4_lstm_for_structure2\\weather_data.xlsx'
wb = load_workbook(filename=filename)
ws = wb.active 
temp1 = []
for row in ws.rows:
    temp2 = []
    for col in row:
        temp2.append(col.value)
    temp1.append(temp2)
temp2 = temp1[1:]
pre_XX_wea = temp2

#load label
filename = 'C:\\Users\\HP\\Desktop\\实验\\流感疫情分析\\工作进展\\2_LSTM预测\\4_lstm_for_structure2\\label.xlsx'
wb = load_workbook(filename=filename)
ws = wb.active 
#load data
temp1 = []
for row in ws.rows:
    for col in row:
        temp2 = col.value
    temp1.append(temp2)
temp2 = temp1[1:]
label = np.array(temp2)

##############################################
#normaliztion
pre_XX = np.array(pre_XX)
pre_XX_wea = np.array(pre_XX_wea)
scaler = MinMaxScaler(feature_range=(0, 1))
pre_XX = scaler.fit_transform(pre_XX)
pre_XX_wea = scaler.fit_transform(pre_XX_wea)

###############################
#split train set and test set
train_XX = []
test_XX = []
train_XX_wea = []
test_XX_wea = []
train_yy = []
test_yy = []

for i_file in range(len(files)):
    train_X = []
    test_X = []
    content = pre_XX[i_file * TOTAL_WEEK : (i_file + 1) * TOTAL_WEEK]
    for i_week in range(TRAIN_WEEK):
        train_X.append(content[i_week : i_week + WEEK_LENGTH])
    for i_week in range(TRAIN_WEEK, TOTAL_WEEK - WEEK_LENGTH):
        test_X.append(content[i_week : i_week + WEEK_LENGTH])
    train_XX.append(np.array(train_X))
    test_XX.append(np.array(test_X))

weather = pre_XX_wea
for i_week in range(TRAIN_WEEK):
    train_XX_wea.append(weather[i_week : i_week + WEEK_LENGTH])
for i_week in range(TRAIN_WEEK, TOTAL_WEEK - WEEK_LENGTH):
    test_XX_wea.append(weather[i_week : i_week + WEEK_LENGTH])
train_XX_wea = np.array(train_XX_wea)
test_XX_wea = np.array(test_XX_wea)

train_yy = label[WEEK_LENGTH:TRAIN_WEEK + WEEK_LENGTH]
test_yy = label[TRAIN_WEEK + WEEK_LENGTH : ]
train_yy = np.array(train_yy)
test_yy = np.array(test_yy)

##############################################
#disorder process
com_XX = [list(train_XX[i]) + list(test_XX[i]) for i in range(AREA_NUM)]
com_XX_wea = list(train_XX_wea) + list(test_XX_wea)
label = list(train_yy) + list(test_yy)

train_index = []
rand_list = [j for j in range(TOTAL_WEEK - WEEK_LENGTH)]
for i in range(TRAIN_WEEK):
    jj = random.choice(rand_list)
    train_index.append(jj)
    rand_list.remove(jj)
test_index = []
for i in range(TOTAL_WEEK - WEEK_LENGTH):
    if i not in train_index:
        test_index.append(i)

#iii = -1
train_XX = [np.array([com_XX[i][k] for k in train_index]) for i in range(AREA_NUM)]
#train_XX = train_XX[0:iii] + train_XX[iii+1:]
test_XX = [np.array([com_XX[i][k] for k in test_index]) for i in range(AREA_NUM)]
#test_XX = test_XX[0:iii] + test_XX[iii+1:]

train_XX_wea = np.array([com_XX_wea[k] for k in train_index])
test_XX_wea = np.array([com_XX_wea[k] for k in test_index])

train_yy = np.array([label[k] for k in train_index])
test_yy = np.array([label[k] for k in test_index])

###############################
# design network
inputs = [Input(shape=(train_XX[0].shape[1], train_XX[0].shape[2])) for i in range(AREA_NUM)]
input_wea = Input(shape=(train_XX_wea.shape[1], train_XX_wea.shape[2]))

hl1s = [LSTM(64)(inputs[i]) for i in range(AREA_NUM)] #32
hl1s_wea = LSTM(64)(input_wea) #32

hl2hs = concatenate(hl1s)

hl3s = concatenate([Dense(16, kernel_regularizer=regularizers.l2(0.01), activity_regularizer=regularizers.l1(0.01))(hl2hs)] + [(hl1s_wea)])
hl4 = Dense(10, kernel_regularizer=regularizers.l2(0.01), activity_regularizer=regularizers.l1(0.01))(hl3s)
hl5 = Dense(1, kernel_regularizer=regularizers.l2(0.01), activity_regularizer=regularizers.l1(0.01))(hl4)

model = Model(inputs = inputs + [input_wea] , output = hl5)
model.compile(loss='mape', optimizer='adam')

# fit network
history = model.fit(train_XX + [train_XX_wea], train_yy, epochs=250, batch_size=32, verbose=1)

#predict
pre = model.predict(test_XX + [test_XX_wea])
pre_yy = pre.reshape(len(pre),)
mape = np.mean(abs((test_yy - pre_yy)/test_yy))
print(mape)
#plt.plot(test_yy, color = 'blue')
#plt.plot(pre, color = 'red')

#disorder prediction
dis_ord_pre = model.predict(train_XX + [train_XX_wea])
dis_ord_pre_yy = dis_ord_pre.reshape(len(dis_ord_pre),)

m = []
for i in range(TOTAL_WEEK - WEEK_LENGTH):
    for j in range(len(train_index)):
        if i == train_index[j]:
            m.append(train_yy[j])
            break
    for j in range(len(test_index)):
        if i == test_index[j]:
            m.append(test_yy[j])
            break
n = []
for i in range(TOTAL_WEEK - WEEK_LENGTH):
    for j in range(len(train_index)):
        if i == train_index[j]:
            n.append(dis_ord_pre_yy[j])
            break
    for j in range(len(test_index)):
        if i == test_index[j]:
            n.append(pre_yy[j])
            break
#plt.plot(m[370:461], color = 'blue')
#plt.plot(n[370:461], color = 'red')

#div 10
p = np.array(m)
q = np.array(n)
#p = scaler.inverse_transform(p)
#q = scaler.inverse_transform(q)
for i in range(len(m)):
    p[i] = p[i] / 10
for i in range(len(n)):
    q[i] = q[i] / 10
plt.plot(p[370:461], color = 'blue')
plt.plot(q[370:461], color = 'red')

np.savetxt('test_yy.csv', p, delimiter = ',')
np.savetxt('pre_yy.csv', q, delimiter = ',')

#test_yy_matrix = np.loadtxt(open("C:\\Users\\HP\Desktop\\实验\\流感疫情分析\\工作进展\\3_对比实验\\2_multi_channel_lstm\\test1\\test_yy.csv","rb"),delimiter=",")
#pre_yy_matrix = np.loadtxt(open("C:\\Users\\HP\Desktop\\实验\\流感疫情分析\\工作进展\\3_对比实验\\2_multi_channel_lstm\\test1\\pre_yy.csv","rb"),delimiter=",")
