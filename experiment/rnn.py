from math import sqrt
from numpy import concatenate
from matplotlib import pyplot
from pandas import read_csv
from pandas import DataFrame
from pandas import concat
from sklearn.preprocessing import MinMaxScaler
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import mean_squared_error
from keras.models import Sequential
from keras.layers import Dense
from keras.layers import LSTM
from keras.layers.recurrent import SimpleRNN
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np
import os
import math
import random

#constant
TOTAL_WEEK = 468
TRAIN_WEEK = 370
WEEK_LENGTH = 6

########################################
#load content data
filename = ''
wb = load_workbook(filename=filename)
ws = wb.active 
temp = []
temp1 = []
for row in ws.rows:
    temp2 = []
    for col in row:
        temp2.append(col.value)
    temp1.append(temp2)
content = temp1[1: TOTAL_WEEK + 1]

#normaliztion
scaler = MinMaxScaler(feature_range=(0, 1))
content = scaler.fit_transform(np.array(content))

#split train set and test set
train_X = []
test_X = []
for i_week in range(TRAIN_WEEK):
    train_X.append(content[i_week : i_week + WEEK_LENGTH])
for i_week in range(TRAIN_WEEK, TOTAL_WEEK - WEEK_LENGTH):
    test_X.append(content[i_week : i_week + WEEK_LENGTH])
train_X = np.array(train_X)
test_X = np.array(test_X)

#####################################
#load label data
filename = ''
wb = load_workbook(filename=filename)
ws = wb.active 
temp = []
temp1 = []
for row in ws.rows:
    for col in row:
        temp2 = col.value
    temp1.append(temp2)
temp = temp1[1:]
label = temp

#split train set and test set
train_y = []
test_y = []
train_y = label[WEEK_LENGTH:TRAIN_WEEK + WEEK_LENGTH]
test_y = label[TRAIN_WEEK + WEEK_LENGTH : ]
train_y = np.array(train_y)
test_y = np.array(test_y)
print(train_X.shape, train_y.shape, test_X.shape, test_y.shape)

##############################################
'''
#disorder process
com_XX = list(train_X) + list(test_X)
label = list(train_y) + list(test_y)

train_index = []
rand_list = [j for j in range(TOTAL_WEEK - WEEK_LENGTH)]
for i in range(TRAIN_WEEK):
    jj = random.choice(rand_list)
    train_index.append(jj)
    rand_list.remove(jj)
test_index = []
for i in range(TOTAL_WEEK - WEEK_LENGTH):
    if i not in train_index:
        test_index.append(i)

train_X = np.array([com_XX[k] for k in train_index])
test_X = np.array([com_XX[k] for k in test_index])

train_y = np.array([label[k] for k in train_index])
test_y = np.array([label[k] for k in test_index])
'''

#################################
# design network
model = Sequential()
model.add(SimpleRNN(256, input_shape=(train_X.shape[1], train_X.shape[2])))  #32
#model.add(Dense(16))
model.add(Dense(1))
model.compile(loss='mape', optimizer='adam')

# fit network
history = model.fit(train_X, train_y, epochs=250, batch_size=32, verbose=1)

#predict
pre = model.predict(test_X)
pre_y = pre.reshape(len(pre),)
mape = np.mean(abs((test_y - pre_y)/test_y))
print(mape)
#plt.plot(test_y, color = 'blue')
#plt.plot(pre, color = 'red')

for i in range(len(test_y)):
    test_y[i] = test_y[i] / 10
for i in range(len(pre_y)):
    pre_y[i] = pre_y[i] / 10
plt.plot(test_y, color = 'blue')
plt.plot(pre_y, color = 'red')

np.savetxt('test_yy.csv', test_y, delimiter = ',')
np.savetxt('pre_yy.csv', pre_y, delimiter = ',')

'''
#disorder prediction
dis_ord_pre = model.predict(train_X)
dis_ord_pre_y = dis_ord_pre.reshape(len(dis_ord_pre),)

m = []
for i in range(TOTAL_WEEK - WEEK_LENGTH):
    for j in range(len(train_index)):
        if i == train_index[j]:
            m.append(train_y[j])
            break
    for j in range(len(test_index)):
        if i == test_index[j]:
            m.append(test_y[j])
            break
n = []
for i in range(TOTAL_WEEK - WEEK_LENGTH):
    for j in range(len(train_index)):
        if i == train_index[j]:
            n.append(dis_ord_pre_y[j])
            break
    for j in range(len(test_index)):
        if i == test_index[j]:
            n.append(pre_y[j])
            break
#plt.plot(m[370:461], color = 'blue')
#plt.plot(n[370:461], color = 'red')

#div 10
p = np.array(m)
q = np.array(n)
#p = scaler.inverse_transform(p)
#q = scaler.inverse_transform(q)
for i in range(len(m)):
    p[i] = p[i] / 10
for i in range(len(n)):
    q[i] = q[i] / 10
plt.plot(p[370:461], color = 'blue')
plt.plot(q[370:461], color = 'red')
'''